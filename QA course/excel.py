from openpyxl import Workbook

workbook = Workbook()
ws = workbook.active
ws.title = "Data"
ws.append(["", "Test1", "Test2", "Test3", "Test4", "Test5", "Test6", "Test7"])
ws.append(["Category", "speakers", "headphones", "mice", "mice", "laptop", "laptop", "tables"])
ws.append(["Product ID", "19", "12", "28", "30", "9", "7", "18"])
ws.append(["Quantity", "2", "2", "", "1", "2", "3"])
ws.append(["COLOR", "RED", "PURPLE", "GREEN", "", "BLUE", "", "YELLOW"])
ws['A1'] = "TZACH"
workbook.save(filename="TestData.xlsx")
#14.1
# with open('story.text', 'a+') as story:
#     story.write('A boy is playing there\n'
#                 'There is a playground\n'
#                 'An airplane is in the sky\n'
#                 'The sky is pink\n'
#                 'Alphabets and numbers are allowed in the passward')

# 14.2
# file = open('story.text', 'r')
# count = 0
# for line in file:
#     if line[0] == 'T':
#         pass
#     else:
#         count += 1
# file.close()
# print(count)

# def The_or_the(text):
#     count = 0
#     for line in text:
#         words = line.split()
#         for word in words:
#             if word == "The":
#                 count += 1
#             elif word == "the":
#                 count += 1
#     return count
# file = open('story.text', 'r')
# print("The number of The or the in story:", The_or_the(file))

# def The_or_the(text):
#     count = 0
#     for line in file:
#         words = line.split()
#         for word in words:
#                 count += 1
#     return count
# file = open('story.text', 'r')
# print("The number of The or the in story:", The_or_the(file))

#14.3:
# def file():
#     with open('story.text', 'r') as story:
#         for line in story:
#             print(line)
# file()
#
# def file():
#     with open('story.text', 'r') as story:
#         count = 0
#         for line in story:
#             count += 1
#         return count
# print(file())

# from openpyxl import Workbook
#
# filename = "hello_world.xlsx"
# workbook = Workbook()
# sheet = workbook.active
#
# sheet["A1"] = "hello"
# sheet["B1"] = "world!"
#
# def print_rows():
#     for row in sheet.iter_rows(values_only=True):
#         print(row)
# sheet["B10"] = "test"
# sheet.insert_cols(idx=1)
# sheet.insert_cols(idx=3, amount=5)
# sheet.delete_cols(idx=3, amount=5)
# sheet.delete_cols(idx=1)
# sheet.insert_rows(idx=1)
# sheet.insert_rows(idx = 3, amount=5)
# print_rows()
# print(workbook.sheetnames)
# workbook.save(filename=filename)


